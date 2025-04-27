import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
from colorama import Fore, Style

from utils_functions import InterfaceStyle


class ExcelWorkbookManager:
    def __init__(self, month_list, filepath="Budget.xlsx"):
        self.filepath = filepath
        self.month_list = month_list
        self.wb = self.file_open_or_creation()


    def file_open_or_creation(self):
        if not os.path.exists(self.filepath):
            wb = Workbook()
            wb.save(self.filepath)
        else:
            wb = load_workbook(self.filepath)
        return wb


    def create_sheet_per_month(self):
        for month in self.month_list:
            if month not in self.wb.sheetnames:
                InterfaceStyle.pretty_print(f"🆕{month} sheet not existing, creation of a new sheet.", top_padding = 0, style = Fore.BLUE)
                self.wb.create_sheet(title=month)
            else:
                InterfaceStyle.pretty_print(f"🆕{month} sheet already exists", top_padding = 0, style = Fore.BLUE)
        self.wb.save(self.filepath)


    def create_recap_sheet_per_month(self, modele_name = "recap_model"):
        if modele_name not in self.wb.sheetnames:
            InterfaceStyle.pretty_print(f"❌ recap modele file not existent - unable to create any monthly recap", top_padding = 0, style = Fore.RED)
            return
        
        for month in self.month_list:
            recap_title = f"{month} - recap"
            
            if recap_title not in self.wb.sheetnames:
                copied_sheet = self.wb.copy_worksheet(self.wb[modele_name])
                copied_sheet.title = recap_title
                InterfaceStyle.pretty_print(f"🆕{recap_title} sheet created", top_padding = 0, style = Fore.GREEN)
            else:
                InterfaceStyle.pretty_print(f"🆕{month} sheet already exists", top_padding = 0, style = Fore.RED)

        self.wb.save(self.filepath)

    def hide_old_sheets(self):
        if len(self.wb.sheetnames) > 2:
            sheet_names_datetime = [datetime.strptime(month, "%B %Y") for month in self.wb.sheetnames].sort()
            sheets_to_hide = sheet_names_datetime[:-2]
            for sheet in sheets_to_hide:
                self.wb[sheet].sheet_state = "hidden"
                self.wb[f"{sheet} - recap"].sheet_state = "hidden"


    def run_file_creation(self):      
        self.create_sheet_per_month()
        self.create_recap_sheet_per_month()
        self.hide_old_sheets()


class DataToExcelManager:
    def __init__(self, transactions, workbook):
        self.df_per_month = transactions.categorize_per_month_as_data_frames()
        self.wb = workbook
        self.month_list = workbook.month_list


    def append_transactions_to_sheet(self):
        for month in self.month_list:
            new_df = self.df_per_month[month]
            existing_df = pd.read_excel(self.wb.filepath, month)

            if not existing_df.empty:
                merged = pd.merge(new_df, existing_df, on=["Narrative", "Amount", "Balance"], how='left', indicator=True)
                df_to_add = merged[merged["_merge"] == "left_only"].drop(columns=["_merge", "Date_y", "Account_y", "Category_y", "Type_y"])
                df_to_add = df_to_add.rename(columns={"Date_x": "Date", "Account_x": "Account", "Category_x": "Category", "Type_x": "Type"})
            else:
                df_to_add = new_df
            
            start_row = len(existing_df)

            with pd.ExcelWriter(self.wb.filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df_to_add.to_excel(writer, sheet_name=month, index=False, startrow=start_row, header = False if start_row > 1 else True)

            InterfaceStyle.pretty_print(f"{month} sheet updated",top_padding = 0, style = Fore.GREEN, )
            


